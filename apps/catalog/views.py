from django.shortcuts import render, get_object_or_404
from rest_framework import viewsets, permissions, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from django_filters.rest_framework import DjangoFilterBackend
import openpyxl
from django.http import HttpResponse
from openpyxl import load_workbook
import io

from .models import Categoria, Curso, Favorito
from .serializers import (
    CategoriaSerializer,
    CursoListSerializer,
    CursoDetailSerializer,
    CursoWriteSerializer,
)

# Roles con acceso de escritura
STAFF_ROLES = {"superadmin", "owner", "admin", "instructor"}
ADMIN_ROLES = {"superadmin", "owner", "admin"}


def _is_staff(user):
    return user.is_authenticated and (user.is_superuser or user.role in STAFF_ROLES)


def _is_admin(user):
    return user.is_authenticated and (user.is_superuser or user.role in ADMIN_ROLES)


class IsAdminOrReadOnly(permissions.BasePermission):
    """Lectura libre. Escritura solo para superadmin/owner/admin/instructor."""
    def has_permission(self, request, view):
        if request.method in permissions.SAFE_METHODS:
            return True
        return _is_staff(request.user)


class CategoriaViewSet(viewsets.ModelViewSet):
    queryset           = Categoria.objects.all()
    serializer_class   = CategoriaSerializer
    permission_classes = [IsAdminOrReadOnly]

    def get_queryset(self):
        qs = super().get_queryset()
        if not _is_admin(self.request.user):
            qs = qs.filter(oculto=False)
        return qs


class CursoViewSet(viewsets.ModelViewSet):
    queryset           = Curso.objects.select_related("categoria").prefetch_related("bloques").all()
    permission_classes = [IsAdminOrReadOnly]
    filter_backends    = [DjangoFilterBackend, filters.SearchFilter, filters.OrderingFilter]
    filterset_fields   = ["categoria__key", "destacado", "mas_vendido", "oculto"]
    search_fields      = ["titulo", "descripcion_corta"]
    ordering_fields    = ["titulo", "creado"]
    ordering           = ["-creado"]
    lookup_field       = "slug"

    def get_serializer_class(self):
        if self.action in ('create', 'update', 'partial_update'):
            return CursoWriteSerializer
        if self.action == 'retrieve':
            return CursoDetailSerializer
        return CursoListSerializer

    def get_queryset(self):
        qs = super().get_queryset()
        if not _is_admin(self.request.user):
            qs = qs.filter(oculto=False)
        return qs

    @action(detail=True, methods=["post"], permission_classes=[permissions.IsAuthenticated])
    def favorito(self, request, slug=None):
        curso = self.get_object()
        fav, created = Favorito.objects.get_or_create(usuario=request.user, curso=curso)
        if not created:
            fav.delete()
            return Response({"favorito": False})
        return Response({"favorito": True})

    @action(detail=False, methods=["get"], permission_classes=[permissions.IsAuthenticated])
    def mis_favoritos(self, request):
        ids    = request.user.favoritos.values_list("curso_id", flat=True)
        cursos = Curso.objects.filter(id__in=ids)
        return Response(CursoListSerializer(cursos, many=True).data)


# ── Vistas HTML ───────────────────────────────────────────────────────────────

def index_page(request):
    return render(request, "index.html")


def curso_page(request, slug):
    curso = get_object_or_404(
        Curso.objects.select_related("categoria").prefetch_related("bloques"),
        slug=slug,
    )
    return render(request, "curso_detalle.html", {"curso": curso})


# ── Excel export / import ─────────────────────────────────────────────────────

def exportar_cursos_excel(request):
    if not _is_admin(request.user):
        return HttpResponse('No autorizado', status=403)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Cursos"
    headers = [
        'titulo', 'slug', 'categoria_key', 'descripcion_corta',
        'emoji', 'imagen_banner', 'imagen_card', 'buy_url',
        'buy_text', 'whatsapp', 'destacado', 'mas_vendido', 'oculto',
    ]
    ws.append(headers)
    for c in Curso.objects.select_related('categoria').all():
        ws.append([
            c.titulo, c.slug,
            c.categoria.key if c.categoria else '',
            c.descripcion_corta, c.emoji,
            c.imagen_banner, c.imagen_card,
            c.buy_url, c.buy_text, c.whatsapp,
            c.destacado, c.mas_vendido, c.oculto,
        ])
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = 'attachment; filename=cursos.xlsx'
    wb.save(response)
    return response


def importar_cursos_excel(request):
    if not _is_admin(request.user):
        return HttpResponse('No autorizado', status=403)
    if request.method != 'POST' or 'archivo' not in request.FILES:
        return HttpResponse('Archivo requerido', status=400)

    archivo = request.FILES['archivo']
    wb      = load_workbook(io.BytesIO(archivo.read()))
    ws      = wb.active
    headers = [cell.value for cell in ws[1]]
    creados = 0
    errores = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        data = dict(zip(headers, row))
        if not data.get('titulo'):
            continue
        try:
            cat = Categoria.objects.filter(key=data.get('categoria_key', '')).first()
            if not cat:
                errores.append(f"Categoría no encontrada: {data.get('categoria_key')}")
                continue
            Curso.objects.update_or_create(
                slug=data.get('slug') or '',
                defaults={
                    'titulo':            data['titulo'],
                    'categoria':         cat,
                    'descripcion_corta': data.get('descripcion_corta', ''),
                    'emoji':             data.get('emoji', '📚'),
                    'imagen_banner':     data.get('imagen_banner', ''),
                    'imagen_card':       data.get('imagen_card', ''),
                    'buy_url':           data.get('buy_url', ''),
                    'buy_text':          data.get('buy_text', 'Comprar ahora'),
                    'whatsapp':          data.get('whatsapp', ''),
                    'destacado':         bool(data.get('destacado', False)),
                    'mas_vendido':       bool(data.get('mas_vendido', False)),
                    'oculto':            bool(data.get('oculto', False)),
                }
            )
            creados += 1
        except Exception as e:
            errores.append(str(e))

    msg = f"Importados: {creados}."
    if errores:
        msg += f" Errores: {'; '.join(errores)}"
    return HttpResponse(msg)